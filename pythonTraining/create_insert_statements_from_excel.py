import openpyxl as px


HEADER_ROW = 3
MIN_ROW = 4
MAX_ROW = 6
MIN_COL = 3
MAX_COL = 7
DST_COL = MAX_COL + 1


def create_insert_statements_from_excel(excel_file_path: str) -> None:
    '''テーブルデータが記述されたExcelファイルからINSERT文を作成する'''
    book = px.load_workbook(excel_file_path)
    for sheetname in book.sheetnames:
        create_insert_statements(book[sheetname])
    book.save(excel_file_path)
    print("Now SQL created.")


def create_insert_statements(sheet: px.worksheet.worksheet.Worksheet) -> None:
    '''ワークシートからINSERT文を作成する'''
    the_former_half = create_the_former_half(sheet)
    table_data = sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL)
    for i, row in enumerate(table_data, start=MIN_ROW):
        statement = the_former_half + " "
        statement += create_the_latter_half(row)
        statement += ";"
        sheet.cell(row=i, column=DST_COL).value = statement
        print(statement)


def create_the_former_half(sheet: px.worksheet.worksheet.Worksheet) -> str:
    '''ワークシートからINSERT文の前半を作成する'''
    table_name = sheet.title
    statement = "INSERT INTO " + table_name + "("
    table_header = list(sheet.iter_rows(
        min_row=HEADER_ROW, max_row=HEADER_ROW, min_col=MIN_COL, max_col=MAX_COL)
    ).pop()
    for cell in table_header:
        delimiter = ", "
        statement += cell.value + delimiter
    statement = statement[0:-1*len(delimiter)] + ")"
    return statement


def create_the_latter_half(row: tuple) -> str:
    '''ワークシートからINSERT文の後半（VALUES句）を作成する'''
    values_phrase = "VALUES("
    for cell in row:
        delimiter = ", "
        values_phrase += parse_cell_value(cell) + delimiter
    values_phrase = values_phrase[0:-1*len(delimiter)] + ")"
    return values_phrase


def parse_cell_value(cell: px.cell.cell.Cell) -> str:
    '''セルの値をVALUE句の値にパースする'''
    if cell.value is None:
        return "''"
    return "'" + str(cell.value) + "'"
